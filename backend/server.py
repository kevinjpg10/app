from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime
import io
import json

# Excel generation
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# PDF generation
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm

# LLM Integration for OCR
from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL')
if not mongo_url:
    raise ValueError("MONGO_URL environment variable is required")
client = AsyncIOMotorClient(mongo_url)
db_name = os.environ.get('DB_NAME')
if not db_name:
    raise ValueError("DB_NAME environment variable is required")
db = client[db_name]

# LLM Key
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')

# Categories for expenses
EXPENSE_CATEGORIES = ["Comida", "Gasolina", "Transporte", "Alojamiento", "Material", "Varios"]

# Create the main app
app = FastAPI(title="Efrain Asistente de Gastos API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============= Models =============

class Expense(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    establishment_name: str = ""
    cif: str = ""
    address: str = ""
    phone: str = ""
    total: float = 0.0
    date: str = ""
    payment_method: str = "efectivo"
    category: str = "Varios"
    image_base64: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    raw_ocr_text: Optional[str] = None

class ExpenseCreate(BaseModel):
    image_base64: Optional[str] = None

class ExpenseManualCreate(BaseModel):
    establishment_name: str
    cif: Optional[str] = ""
    address: Optional[str] = ""
    phone: Optional[str] = ""
    total: float = 0.0
    date: Optional[str] = ""
    payment_method: str = "efectivo"
    category: str = "Varios"

class ExpenseUpdate(BaseModel):
    establishment_name: Optional[str] = None
    cif: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    total: Optional[float] = None
    date: Optional[str] = None
    payment_method: Optional[str] = None
    category: Optional[str] = None

class ExpenseResponse(BaseModel):
    id: str
    establishment_name: str
    cif: str
    address: str
    phone: str
    total: float
    date: str
    payment_method: str
    category: str = "Varios"
    image_base64: Optional[str] = None
    created_at: datetime

# ============= OCR Function =============

async def extract_ticket_data(image_base64: str) -> dict:
    """Use GPT-4o to extract data from a ticket image"""
    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"ticket-ocr-{uuid.uuid4()}",
            system_message="""Eres un experto en extraer datos de tickets y facturas españolas.
            Tu tarea es analizar la imagen del ticket y extraer la siguiente información:
            - Nombre del establecimiento
            - CIF (número de identificación fiscal)
            - Dirección
            - Teléfono
            - Total pagado (solo el número)
            - Fecha (en formato DD/MM/YYYY)
            - Método de pago (determinar si fue "tarjeta" o "efectivo")
            - Categoría del gasto (una de: "Comida", "Gasolina", "Transporte", "Alojamiento", "Material", "Varios")
            
            Para determinar la categoría:
            - Comida: restaurantes, cafeterías, supermercados, bares, fast food
            - Gasolina: gasolineras, estaciones de servicio
            - Transporte: taxis, uber, autobuses, trenes, aviones, parking
            - Alojamiento: hoteles, hostales, airbnb
            - Material: papelerías, ferreterías, tiendas de suministros
            - Varios: todo lo demás
            
            Responde SIEMPRE en formato JSON con estas claves exactas:
            {
                "establishment_name": "nombre",
                "cif": "número CIF",
                "address": "dirección completa",
                "phone": "teléfono",
                "total": 0.00,
                "date": "DD/MM/YYYY",
                "payment_method": "tarjeta" o "efectivo",
                "category": "categoría"
            }
            
            Si no puedes encontrar algún dato, usa una cadena vacía para textos o 0.0 para números.
            """
        ).with_model("openai", "gpt-4o")
        
        image_content = ImageContent(image_base64=image_base64)
        user_message = UserMessage(
            text="Por favor, extrae todos los datos de este ticket. Responde solo con el JSON.",
            file_contents=[image_content]
        )
        
        response = await chat.send_message(user_message)
        logger.info(f"OCR Response: {response}")
        
        clean_response = response.strip()
        if clean_response.startswith("```"):
            clean_response = clean_response.split("```")[1]
            if clean_response.startswith("json"):
                clean_response = clean_response[4:]
        clean_response = clean_response.strip()
        
        data = json.loads(clean_response)
        return data
        
    except Exception as e:
        logger.error(f"Error extracting ticket data: {e}")
        return {
            "establishment_name": "",
            "cif": "",
            "address": "",
            "phone": "",
            "total": 0.0,
            "date": "",
            "payment_method": "efectivo",
            "category": "Varios"
        }

# ============= Routes =============

@api_router.get("/")
async def root():
    return {"message": "Efrain Asistente de Gastos API", "version": "1.0.0"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "service": "efrain-gastos"}

@api_router.get("/categories")
async def get_categories():
    return {"categories": EXPENSE_CATEGORIES}

@api_router.post("/expenses", response_model=ExpenseResponse)
async def create_expense(expense_data: ExpenseCreate):
    try:
        if expense_data.image_base64 and len(expense_data.image_base64) > 100:
            extracted_data = await extract_ticket_data(expense_data.image_base64)
        else:
            extracted_data = {
                "establishment_name": "", "cif": "", "address": "", "phone": "",
                "total": 0.0, "date": "", "payment_method": "efectivo", "category": "Varios"
            }
        
        expense = Expense(
            establishment_name=extracted_data.get("establishment_name", ""),
            cif=extracted_data.get("cif", ""),
            address=extracted_data.get("address", ""),
            phone=extracted_data.get("phone", ""),
            total=float(extracted_data.get("total", 0.0)),
            date=extracted_data.get("date", ""),
            payment_method=extracted_data.get("payment_method", "efectivo"),
            category=extracted_data.get("category", "Varios"),
            image_base64=expense_data.image_base64 if expense_data.image_base64 and len(expense_data.image_base64) > 100 else None,
            raw_ocr_text=str(extracted_data) if expense_data.image_base64 and len(expense_data.image_base64) > 100 else None
        )
        
        expense_dict = expense.model_dump()
        await db.expenses.insert_one(expense_dict)
        return ExpenseResponse(**expense_dict)
        
    except Exception as e:
        logger.error(f"Error creating expense: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/expenses/manual", response_model=ExpenseResponse)
async def create_manual_expense(expense_data: ExpenseManualCreate):
    try:
        expense = Expense(
            establishment_name=expense_data.establishment_name,
            cif=expense_data.cif or "",
            address=expense_data.address or "",
            phone=expense_data.phone or "",
            total=expense_data.total,
            date=expense_data.date or "",
            payment_method=expense_data.payment_method,
            category=expense_data.category,
            image_base64=None,
            raw_ocr_text=None
        )
        
        expense_dict = expense.model_dump()
        await db.expenses.insert_one(expense_dict)
        return ExpenseResponse(**expense_dict)
        
    except Exception as e:
        logger.error(f"Error creating manual expense: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/expenses", response_model=List[ExpenseResponse])
async def get_expenses():
    # Projection to fetch only needed fields, excluding large image_base64 for list view
    projection = {
        "id": 1, "establishment_name": 1, "cif": 1, "address": 1, "phone": 1,
        "total": 1, "date": 1, "payment_method": 1, "category": 1, "created_at": 1,
        "image_base64": 1, "_id": 0
    }
    expenses = await db.expenses.find({}, projection).sort("created_at", -1).to_list(1000)
    return [ExpenseResponse(**expense) for expense in expenses]

@api_router.get("/expenses/{expense_id}", response_model=ExpenseResponse)
async def get_expense(expense_id: str):
    expense = await db.expenses.find_one({"id": expense_id})
    if not expense:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    return ExpenseResponse(**expense)

@api_router.put("/expenses/{expense_id}", response_model=ExpenseResponse)
async def update_expense(expense_id: str, update_data: ExpenseUpdate):
    expense = await db.expenses.find_one({"id": expense_id})
    if not expense:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if update_dict:
        await db.expenses.update_one({"id": expense_id}, {"$set": update_dict})
    
    updated_expense = await db.expenses.find_one({"id": expense_id})
    return ExpenseResponse(**updated_expense)

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str):
    result = await db.expenses.delete_one({"id": expense_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    return {"message": "Gasto eliminado correctamente"}

@api_router.delete("/expenses")
async def delete_all_expenses():
    result = await db.expenses.delete_many({})
    return {"message": f"Se eliminaron {result.deleted_count} gastos", "deleted_count": result.deleted_count}

@api_router.patch("/expenses/{expense_id}/payment-method")
async def toggle_payment_method(expense_id: str):
    expense = await db.expenses.find_one({"id": expense_id})
    if not expense:
        raise HTTPException(status_code=404, detail="Gasto no encontrado")
    
    current_method = expense.get("payment_method", "efectivo").lower()
    new_method = "tarjeta" if current_method == "efectivo" else "efectivo"
    
    await db.expenses.update_one({"id": expense_id}, {"$set": {"payment_method": new_method}})
    updated_expense = await db.expenses.find_one({"id": expense_id})
    return ExpenseResponse(**updated_expense)

@api_router.get("/expenses/summary/stats")
async def get_expenses_summary():
    # Projection to fetch only needed fields for calculations
    projection = {"total": 1, "payment_method": 1, "category": 1, "_id": 0}
    expenses = await db.expenses.find({}, projection).to_list(1000)
    
    total_general = sum(e.get("total", 0) for e in expenses)
    total_tarjeta = sum(e.get("total", 0) for e in expenses if e.get("payment_method", "").lower() == "tarjeta")
    total_efectivo = sum(e.get("total", 0) for e in expenses if e.get("payment_method", "").lower() != "tarjeta")
    
    count_tarjeta = len([e for e in expenses if e.get("payment_method", "").lower() == "tarjeta"])
    count_efectivo = len(expenses) - count_tarjeta
    
    categories_data = {}
    for cat in EXPENSE_CATEGORIES:
        cat_expenses = [e for e in expenses if e.get("category", "Varios") == cat]
        cat_total = sum(e.get("total", 0) for e in cat_expenses)
        categories_data[cat] = {
            "count": len(cat_expenses),
            "total": round(cat_total, 2),
            "percentage": round((cat_total / total_general * 100) if total_general > 0 else 0, 1)
        }
    
    return {
        "total_expenses": len(expenses),
        "total_amount": round(total_general, 2),
        "card_payments": {"count": count_tarjeta, "total": round(total_tarjeta, 2)},
        "cash_payments": {"count": count_efectivo, "total": round(total_efectivo, 2)},
        "categories": categories_data
    }

@api_router.get("/expenses/export/excel")
async def export_expenses_excel():
    # Exclude image_base64 for better performance
    projection = {
        "establishment_name": 1, "cif": 1, "address": 1, "phone": 1,
        "total": 1, "date": 1, "category": 1, "payment_method": 1, "created_at": 1, "_id": 0
    }
    expenses = await db.expenses.find({}, projection).sort("created_at", -1).to_list(1000)
    
    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "Todos los Gastos"
    ws_tarjeta = wb.create_sheet("Pagos con Tarjeta")
    ws_efectivo = wb.create_sheet("Pagos en Efectivo")
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ["Establecimiento", "CIF", "Dirección", "Teléfono", "Total (€)", "Fecha", "Categoría", "Método de Pago"]
    
    def setup_sheet(ws):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
    
    setup_sheet(ws_all)
    setup_sheet(ws_tarjeta)
    setup_sheet(ws_efectivo)
    
    row_all, row_tarjeta, row_efectivo = 2, 2, 2
    total_all, total_tarjeta, total_efectivo = 0, 0, 0
    
    for expense in expenses:
        data = [
            expense.get("establishment_name", ""), expense.get("cif", ""),
            expense.get("address", ""), expense.get("phone", ""),
            expense.get("total", 0), expense.get("date", ""),
            expense.get("category", "Varios"), expense.get("payment_method", "efectivo")
        ]
        
        for col, value in enumerate(data, 1):
            ws_all.cell(row=row_all, column=col, value=value).border = thin_border
        row_all += 1
        total_all += expense.get("total", 0)
        
        if expense.get("payment_method", "").lower() == "tarjeta":
            for col, value in enumerate(data, 1):
                ws_tarjeta.cell(row=row_tarjeta, column=col, value=value).border = thin_border
            row_tarjeta += 1
            total_tarjeta += expense.get("total", 0)
        else:
            for col, value in enumerate(data, 1):
                ws_efectivo.cell(row=row_efectivo, column=col, value=value).border = thin_border
            row_efectivo += 1
            total_efectivo += expense.get("total", 0)
    
    def add_total_row(ws, row, total):
        ws.cell(row=row, column=4, value="TOTAL:").font = Font(bold=True)
        total_cell = ws.cell(row=row, column=5, value=total)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '€#,##0.00'
    
    add_total_row(ws_all, row_all, total_all)
    add_total_row(ws_tarjeta, row_tarjeta, total_tarjeta)
    add_total_row(ws_efectivo, row_efectivo, total_efectivo)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"gastos_efrain_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename={filename}"})

@api_router.get("/expenses/export/pdf")
async def export_expenses_pdf():
    # Projection to fetch only needed fields for PDF report
    projection = {
        "establishment_name": 1, "cif": 1, "address": 1, "phone": 1,
        "total": 1, "date": 1, "category": 1, "payment_method": 1, "created_at": 1, "_id": 0
    }
    expenses = await db.expenses.find({}, projection).sort("created_at", -1).to_list(1000)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24,
                                 textColor=colors.HexColor('#4A90D9'), spaceAfter=30, alignment=1)
    subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Normal'], fontSize=12,
                                    textColor=colors.HexColor('#666666'), spaceAfter=20, alignment=1)
    section_style = ParagraphStyle('SectionTitle', parent=styles['Heading2'], fontSize=14,
                                   textColor=colors.HexColor('#333333'), spaceBefore=20, spaceAfter=10)
    
    elements = []
    elements.append(Paragraph("EFRAIN - Informe de Gastos", title_style))
    elements.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 20))
    
    total_general = sum(e.get("total", 0) for e in expenses)
    total_tarjeta = sum(e.get("total", 0) for e in expenses if e.get("payment_method", "").lower() == "tarjeta")
    total_efectivo = total_general - total_tarjeta
    
    elements.append(Paragraph("RESUMEN", section_style))
    summary_data = [
        ["Concepto", "Cantidad", "Importe"],
        ["Total de gastos", str(len(expenses)), f"€{total_general:.2f}"],
        ["Pagos con tarjeta", str(len([e for e in expenses if e.get("payment_method", "").lower() == "tarjeta"])), f"€{total_tarjeta:.2f}"],
        ["Pagos en efectivo", str(len([e for e in expenses if e.get("payment_method", "").lower() != "tarjeta"])), f"€{total_efectivo:.2f}"],
    ]
    
    summary_table = Table(summary_data, colWidths=[8*cm, 4*cm, 4*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A90D9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#DDDDDD')),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    elements.append(Paragraph("POR CATEGORÍAS", section_style))
    cat_data = [["Categoría", "Gastos", "Total", "%"]]
    for cat in EXPENSE_CATEGORIES:
        cat_expenses = [e for e in expenses if e.get("category", "Varios") == cat]
        cat_total = sum(e.get("total", 0) for e in cat_expenses)
        cat_pct = (cat_total / total_general * 100) if total_general > 0 else 0
        if len(cat_expenses) > 0:
            cat_data.append([cat, str(len(cat_expenses)), f"€{cat_total:.2f}", f"{cat_pct:.1f}%"])
    
    if len(cat_data) > 1:
        cat_table = Table(cat_data, colWidths=[5*cm, 3*cm, 4*cm, 3*cm])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#DDDDDD')),
        ]))
        elements.append(cat_table)
    elements.append(Spacer(1, 20))
    
    elements.append(Paragraph("DETALLE DE GASTOS", section_style))
    if expenses:
        expense_data = [["Fecha", "Establecimiento", "Categoría", "Método", "Total"]]
        for exp in expenses:
            expense_data.append([
                exp.get("date", "Sin fecha"), exp.get("establishment_name", "Sin nombre")[:25],
                exp.get("category", "Varios"), exp.get("payment_method", "efectivo").capitalize(),
                f"€{exp.get('total', 0):.2f}"
            ])
        
        expense_table = Table(expense_data, colWidths=[2.5*cm, 6*cm, 2.5*cm, 2.5*cm, 2.5*cm])
        expense_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9B59B6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
        ]))
        elements.append(expense_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"informe_efrain_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf",
                            headers={"Content-Disposition": f"attachment; filename={filename}"})

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
